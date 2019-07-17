import openpyxl as xl  # To create alias and to make code shorter
from openpyxl.chart import BarChart, Reference  # To import BarChart & Reference

#-------------------------------------------------------------------------------------------------#
# A function to find a max number from the given list
#-------------------------------------------------------------------------------------------------#


def find_max_from_list(numbers):
    max_number = 0
    for number in numbers:
        if number > max_number:
            max_number = number
    return max_number

#*************************************************************************************************#
# Pypi & Pip (Huh ?) : https://pypi.org/ (Find, install and publish Python packages with the Python Package Index)
#*************************************************************************************************#

# To import new charts from openpyxl

# How to install packages from the PYPI.org ?
# openpyxl : A python package to work with excel spreadsheets
# A command to install the above package : 'pip install openpyxl'

#-------------------------------------------------------------------------------------------------#
# A function to open spreadsheet, calculates operations and
# update the spreadsheets and also add charts to those files
#-------------------------------------------------------------------------------------------------#


def process_workbook(filename):
    # A statement to create an object of a file
    work_book = xl.load_workbook(filename)
    # Access a specific sheet in a spreadsheet and this will return that sheet
    sheet1 = work_book["Sheet1"]

    # cell1 = sheet1["A1"] # By providing coordinates in a sheet, user can have an access to a cell
    # cell1 = sheet1.cell(1, 1) # Another method to get cell from a sheet by passing row & column number
    # print(cell1.value) # This will print the exact value that cell contains
    # print(sheet1.max_row) # This will print the exact number of rown in a specific sheet

    for row in range(sheet1.max_row):
        row += 1
        # print(row)
        if row != 1:
            cell = sheet1.cell(row, 3)
            print(cell.value)
            updated_price = cell.value * 0.9
            print(updated_price)
            # To create a new cell called a column # 4 for each row
            updated_cell = sheet1.cell(row, 4)
            updated_cell.value = updated_price

    row_values = Reference(sheet1,
                           min_row=2,
                           max_row=sheet1.max_row,
                           min_col=4,
                           max_col=4)  # To select row == 2,3,4 in a selected sheet

    chart = BarChart()  # To create an instance of a BarChart
    chart.add_data(row_values)  # pass arguments to chart
    sheet1.add_chart(chart, 'E2')

    # work_book.save("transactions2.xlsx") # A statement to save these changes in new .xlsx file in a project.
    # A statement to override & save the changes within a same file.
    work_book.save(filename)
